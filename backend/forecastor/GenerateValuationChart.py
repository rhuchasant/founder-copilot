import boto3
import io
import json
import uuid
import base64
import matplotlib.pyplot as plt
import csv
from openpyxl import load_workbook

s3 = boto3.client('s3')
bucket_name = "startup-valuation-graphs"

def lambda_handler(event, context):
    try:
        print("Incoming event:", event)

        # Parse base64 file from request
        body = json.loads(event["body"])
        file_content = base64.b64decode(body["file"])
        filename = body.get("filename", "uploaded.csv")

        months = []
        revenue = []
        burn_rate = []

        if filename.endswith(".csv"):
            decoded = io.StringIO(file_content.decode("utf-8"))
            reader = csv.DictReader(decoded)
            for row in reader:
                months.append(row.get("Month"))
                revenue.append(float(row.get("Revenue", 0)))
                burn_rate.append(float(row.get("Burn Rate", 0)))

        elif filename.endswith(".xlsx"):
            wb = load_workbook(io.BytesIO(file_content))
            sheet = wb.active
            headers = list(next(sheet.iter_rows(values_only=True)))
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                months.append(str(row_dict.get("Month")))
                revenue.append(float(row_dict.get("Revenue", 0)))
                burn_rate.append(float(row_dict.get("Burn Rate", 0)))

        else:
            return {
                "statusCode": 400,
                "headers": {
                    "Content-Type": "application/json",
                    "Access-Control-Allow-Origin": "*"
                },
                "body": json.dumps({"error": "Unsupported file type. Use .csv or .xlsx."})
            }

        if not months or not revenue:
            return {
                "statusCode": 400,
                "headers": {
                    "Content-Type": "application/json",
                    "Access-Control-Allow-Origin": "*"
                },
                "body": json.dumps({"error": "Missing required data in file."})
            }

        # Calculate expenses (Expenses = Revenue + Burn Rate)
        expenses = [rev + burn for rev, burn in zip(revenue, burn_rate)]

        # Calculate net profit and cumulative cash
        net_profit = [rev - exp for rev, exp in zip(revenue, expenses)]
        cumulative_cash = [net_profit[0] if net_profit else 0]
        for i in range(1, len(net_profit)):
            cumulative_cash.append(cumulative_cash[i - 1] + net_profit[i])

        # Find break-even months (net profit >= 0)
        break_even_months = [i for i, np in enumerate(net_profit) if np >= 0]

        # Plotting
        plt.figure(figsize=(14, 7))

        plt.plot(months, revenue, label="Revenue", marker="o")
        plt.plot(months, expenses, label="Expenses", marker="o")
        plt.plot(months, burn_rate, label="Burn Rate", marker="o")
        plt.plot(months, net_profit, label="Net Profit", marker="o")

        plt.plot(months, cumulative_cash, label="Cash Runway", marker="o", color="purple")
        plt.fill_between(months, cumulative_cash, color="purple", alpha=0.15)

        # Shade break-even months
        for i in break_even_months:
            plt.axvspan(i - 0.4, i + 0.4, color='green', alpha=0.1)

        # Annotations
        max_rev_index = revenue.index(max(revenue))
        plt.annotate("Highest Revenue", (max_rev_index, revenue[max_rev_index]),
                     textcoords="offset points", xytext=(0,10), ha='center', color='blue')

        if break_even_months:
            first_be_index = break_even_months[0]
            plt.annotate("Break-even Start", (first_be_index, net_profit[first_be_index]),
                         textcoords="offset points", xytext=(0,-15), ha='center', color='green')

        plt.title("Advanced Startup Financial Forecast")
        plt.xlabel("Month")
        plt.ylabel("Amount ($)")
        plt.grid(True)
        plt.legend()
        plt.tight_layout()

        # Save chart to buffer
        buf = io.BytesIO()
        plt.savefig(buf, format="png", bbox_inches="tight")
        buf.seek(0)

        # Upload chart to S3
        chart_filename = f"forecast_{uuid.uuid4()}.png"
        s3.upload_fileobj(
            buf,
            bucket_name,
            chart_filename,
            ExtraArgs={"ContentType": "image/png"}
        )

        image_url = f"https://{bucket_name}.s3.amazonaws.com/{chart_filename}"

        return {
            "statusCode": 200,
            "headers": {
                "Content-Type": "application/json",
                "Access-Control-Allow-Origin": "*"
            },
            "body": json.dumps({"graph_url": image_url})
        }

    except Exception as e:
        print("Error occurred:", str(e))
        return {
            "statusCode": 500,
            "headers": {
                "Content-Type": "application/json",
                "Access-Control-Allow-Origin": "*"
            },
            "body": json.dumps({"error": str(e)})
        }
